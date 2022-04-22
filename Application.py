import pprint
from datetime import datetime
import openpyxl

#setting up the language
wkbk = openpyxl.load_workbook("is352 language code.xlsx")
sheet = wkbk.active
langDict = {}
for line in range(2, sheet.max_row):
    code = sheet.cell(row=line, column=1).value
    lang = sheet.cell(row=line, column=2).value
    langDict[code] = lang
langDict['cn'] = 'Chinese'

# read watchlist
# search function
Application = None


def printMenu():
    print(f"Welcome to {Application}!")
    print("Would you like to: \n1. Search for a movie or show\n2. Go to your watchlist\n3. Exit")
    userInput = input("Please enter a number:  ").strip()
    if userInput != "1" and userInput != "2" and userInput != "3":
        print("That option was not provided, please choose again. ")
        userInput = printMenu()
    return userInput


def searchMenu():

    print("What would you like to search for?")
    print("1. Title \n2. Runtime \n3. Release Date \n4. Language")
    toSearch = input("Enter option:")
    if toSearch == '1':
        title = input("Enter title: ")
        toPrint = search(title, 't', movieDict, showDict)
    elif toSearch == '2':
        # COME BACK TO THIS
        # DO WE WANT TO HAVE A RUNTIME RANGE
        print("Would you like a movie that is: ")
        print("1. Less then an hour\n2. In between 1 and 2 hours\n3. Longer then 2 hours")
        runtime = input("Enter: ")
        toPrint = search(runtime, 'r', movieDict, showDict)
    elif toSearch == '3':
        print("Select the range the release year falls within : ")
        print("1. Before 1920\n2. After 1920 and before 1950 \n3. The 1950's\n4. The 1960's\n5. The 1970's\n6. The 1980's\n7. The 1990's\n8. The 2000's\n9. The 2010's\n10. The 2020's")
        date = input("Enter: ")
        toPrint = search(date, 'd', movieDict, showDict)
    elif toSearch == '4':
        lang = input("Enter language: ")
        toPrint = search(lang, 'l', movieDict, showDict)
    else:
        print("That was not an option. Please enter a new command.")
        searchMenu()
    print(toPrint.keys())


def search(input, flag, mDict, sDict): #flag can be t, r, d
    results = {}
    if flag == 't':
        #search by title
        for title in mDict.keys():
            if input.upper() in title.upper():
                # good
                results[title] = mDict[title]
        for title in sDict.keys():
            if input.upper() in title.upper():
                results[title] = mDict[title]
    elif flag == 'r':
        # search by runtime
        for item in movieDict:
            if input == "1" and mDict[item]['runtime'] < 60:
                results[item] = mDict[item]
            elif input == "2" and (mDict[item]['runtime'] >= 60 and mDict[item]['runtime'] < 120):
                results[item] = mDict[item]
            elif input == '3' and mDict[item]['runtime'] > 120:
                results[item] = mDict[item]
    elif flag == 'd':
        for item in movieDict:
            if input == "1" and mDict[item]['releaseDate'] < 2015:
                results[item] = mDict[item]
            elif input == "2" and (mDict[item]['releaseDate'] >= 1920 and mDict[item]['releaseDate'] < 1950):
                results[item] = mDict[item]
            elif input == "3" and (mDict[item]['releaseDate'] >= 1950 and mDict[item]['releaseDate'] < 1960):
                results[item] = mDict[item]
            elif input == "4" and (mDict[item]['releaseDate'] >= 1960 and mDict[item]['releaseDate'] < 1970):
                results[item] = mDict[item]
            elif input == "5" and (mDict[item]['releaseDate'] >= 1970 and mDict[item]['releaseDate'] < 1980):
                results[item] = mDict[item]
            elif input == "6" and (mDict[item]['releaseDate'] >= 1980 and mDict[item]['releaseDate'] < 1990):
                results[item] = mDict[item]
            elif input == "7" and (mDict[item]['releaseDate'] >= 1990 and mDict[item]['releaseDate'] < 2000):
                results[item] = mDict[item]
            elif input == "8" and (mDict[item]['releaseDate'] >= 2000 and mDict[item]['releaseDate'] < 2010):
                results[item] = mDict[item]
            elif input == "9" and (mDict[item]['releaseDate'] >= 2010 and mDict[item]['releaseDate'] < 2020):
                results[item] = mDict[item]
            elif input == "10" and mDict[item]['releaseDate'] >= 2020:
                results[item] = mDict[item]

        #search by release Date
    elif flag == 'l':
        # search by language
        for item in mDict:
            if input.lower() == langDict[mDict[item]['language']].lower():
                results[item] = mDict[item]
        for item in sDict:
            if input.lower() == langDict[sDict[item]['language']].lower():
                results[item] = sDict[item]
    return results


# def viewWatchlist():
#     for item in watchlist:
#         print(item)

# def addWatchlist():

# def writeWatchlist():
#     # write at the end of running

# read in the data, into a dict.
# shows
fileName = "Show_Data.txt"
with open(fileName) as filePointer:
    showFileContents = filePointer.read().strip()
    shows = showFileContents.split('\n')
    showDict = {}

for show in shows:
    elements = show.split('\t')
    releaseDate = elements[1][-4:].strip()
    if releaseDate != 'None':
        releaseDate = int(releaseDate)
    else:
        releaseDate = 0
    toAdd = {'releaseDate' : releaseDate, 'misc' : elements[2], 'language' : elements[3]}
    if elements[0] not in showDict:
        showDict[elements[0]] = toAdd


# movies
fileName = "Movie_Data.txt"
with open(fileName) as filePointer:
    movieFileContents = filePointer.read().strip()
    movies = movieFileContents.split('\n')
    movieDict = {}

for movie in movies:
    elements = movie.split('\t')
    releaseDate = elements[1][-4:].strip()
    runtime = elements[2].strip()
    if releaseDate != 'None':
        releaseDate = int(releaseDate)
    else:
        releaseDate = 0
    if runtime != 'None':
        runtime = int(runtime)
    else:
        runtime = 0
    toAdd = {'releaseDate' : releaseDate, 'runtime' : runtime, 'language' : elements[3]}
    if elements[0] not in movieDict:
        movieDict[elements[0]] = toAdd


# watchlist
# fileName = "Watchlist.txt"
# with open(fileName, 'w') as filePointer:
#     watchlist = filePointer.read().strip().split('\n')

# aka main
EXIT = "3"
userInput = printMenu()
while True:
    if userInput == EXIT:

        break
    if userInput == "1":
        # search for the movie with given parameters
        # ask for them
        searchMenu()
    if userInput == "2":
        # show them their watchlist
        print("got to watchlist")
    userInput = printMenu()

