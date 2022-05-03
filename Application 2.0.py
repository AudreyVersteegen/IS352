import pprint
import openpyxl
# read watchlist
Application = "Pandora's Box"


def printMenu():
    print(f"Welcome to {Application}!")
    print("Would you like to: \n1. Search for a movie or show\n2. Go to your watchlist\n3. Exit")
    userInput = input("Please enter a number:  ").strip()
    if userInput != "1" and userInput != "2" and userInput != "3":
        print("That option was not provided, please choose again. ")
        userInput = printMenu()
    return userInput

# search functions
def searchMenu():
    mDict = mediaDict
    searching = True
    while searching:
        print("What would you like to search for?")
        print("1. Title \n2. Runtime \n3. Release Date \n4. Language\n5. Return to main menu")
        toSearch = input("Enter option:")
        if toSearch == '1':
            title = input("Enter title: ")
            mDict = search(title, 't', mDict)
        elif toSearch == '2':
            # COME BACK TO THIS
            # DO WE WANT TO HAVE A RUNTIME RANGE
            print("Would you like a movie that is: ")
            print("1. Less then an hour\n2. In between 1 and 2 hours\n3. Longer then 2 hours")
            runtime = input("Enter: ")
            mDict = search(runtime, 'r', mDict)
        elif toSearch == '3':
            print("Select the range the release year falls within : ")
            print("1. Before 1920\n2. After 1920 and before 1950 \n3. The 1950's\n4. The 1960's\n"
                  "5. The 1970's\n6. The 1980's\n7. The 1990's\n8. The 2000's\n9. The 2010's\n10. The 2020's")
            date = input("Enter: ")
            mDict = search(date, 'd', mDict)
        elif toSearch == '4':
            lang = input("Enter language: ")
            mDict = search(lang, 'l', mDict)
        elif toSearch == '5':
            break
        else:
            print("That was not an option. Please enter a new command.")
        if len(mDict) > 0:
            for item in mDict:
                print(item + "\tReleased: " + str(mDict[item]['releaseDate']))
            yn = input('Would you like to narrow down these results? [y/n] ').strip().lower()
            if yn != 'y':
                searching = False
                wlQuestion = input("Would you like to add any of these to your watchlist? [y/n]")
                if wlQuestion.strip() == "y":
                    finalAsk = input("Enter the title of the item you would like to add: ")
                    for item in mDict.keys():
                        if finalAsk.lower() == item.lower():
                            print("Item added to watchlist!")
                            addWatchlist(watchlist, item)
        else:
            print("There are no films or movies that meet those qualifications.")


def search(input, flag, mDict):  # flag can be t, r, d
    results = {}
    input = input.strip()
    if flag == 't':
        # search by title
        for title in mDict.keys():
            if input.upper() in title.upper():
                results[title] = mDict[title]
    elif flag == 'r':
        # search by runtime
        for item in mDict:
            if 'runtime' in mDict[item]:
                if input == "1" and mDict[item]['runtime'] < 60:
                    results[item] = mDict[item]
                elif input == "2" and (mDict[item]['runtime'] >= 60 and mDict[item]['runtime'] < 120):
                    results[item] = mDict[item]
                elif input == '3' and mDict[item]['runtime'] > 120:
                    results[item] = mDict[item]
    elif flag == 'd':
        # search by release date
        eras = [0, 1920, 1950, 1960, 1970, 1980, 1990, 2000, 2010, 2020, 2030]
        if input in ["1", "2", "3", '4', '5', '6', '7', '8', '9', '10']:
            for item in mDict:
                start = eras[int(input) - 1]
                end = eras[int(input)]
                date = mDict[item]['releaseDate']
                if date in range(start, end):
                    results[item] = mDict[item]
    elif flag == 'l':
        # search by language
        for item in mDict:
            if input.lower() == langDict[mDict[item]['language']].lower():
                results[item] = mDict[item]
    return results


def viewWatchlist(watchlist):
    print('Your watchlist: ')
    wlPos = 0
    for item in watchlist:
        wlPos += 1
        print(f'{wlPos}. {item}')
    userInput = input("Would you like to: \n1. Delete an item\n2. Return to main menu \n Enter number:").strip()
    if userInput == '1':
        # which item do you want to ditch
        toDelete = int(input("Enter the number of the movie you would like to remove: "))
        # delete it
        watchlist.pop(toDelete - 1)
        viewWatchlist(watchlist)
    elif userInput != '2':
        print("That option was not provided, please choose again. ")
        viewWatchlist(watchlist)


def addWatchlist(watchlist, item):
    if item not in watchlist:
        watchlist.append(item)

def writeWatchlist(watchlist, fileName):
    # write at the end of running
    with open(fileName, 'w') as fileHandler:
        for item in watchlist:
            fileHandler.write(item + '\n')

# setting up the language
wkbk = openpyxl.load_workbook("is352_language_code.xlsx")
sheet = wkbk.active
langDict = {}
for line in range(2, sheet.max_row):
    code = sheet.cell(row=line, column=1).value
    lang = sheet.cell(row=line, column=2).value
    langDict[code] = lang
langDict['cn'] = 'Chinese'

# read in the data, into a dict.
# shows
fileName = "Show_Data.txt"
mediaDict = {}
with open(fileName) as filePointer:
    showFileContents = filePointer.read().strip()
    shows = showFileContents.split('\n')

for show in shows:
    elements = show.split('\t')
    releaseDate = elements[1][-4:].strip()
    if releaseDate != 'None':
        releaseDate = int(releaseDate)
    else:
        releaseDate = 0
    toAdd = {'releaseDate': releaseDate, 'misc': elements[2], 'language': elements[3]}
    if elements[0] not in mediaDict:
        mediaDict[elements[0]] = toAdd

# movies
fileName = "Movie_Data.txt"
with open(fileName) as filePointer:
    movieFileContents = filePointer.read().strip()
    movies = movieFileContents.split('\n')

for movie in movies:
    elements = movie.split('\t')
    releaseDate = elements[1][-4:].strip()
    runtime = elements[2].strip()
    if releaseDate != 'None':
        releaseDate = int(releaseDate)
    else:
        releaseDate = 0
    if runtime != 'None':
        runtime = int(runtime)
    else:
        runtime = 0
    toAdd = {'releaseDate': releaseDate, 'runtime': runtime, 'language': elements[3]}
    if elements[0] not in mediaDict:
        mediaDict[elements[0]] = toAdd

# watchlist
fileName = "Watchlist.txt"
with open(fileName, 'r+') as filePointer:
    wl = filePointer.read()
    watchlist = wl.strip().split('\n')


# aka main
EXIT = "3"
userInput = printMenu()
while True:
    if userInput == EXIT:
        break
    if userInput == "1":
        # search for the movie with given parameters
        # ask for them
        searchMenu()
    if userInput == "2":
        # show them their watchlist
        viewWatchlist(watchlist)
    userInput = printMenu()
writeWatchlist(watchlist, fileName)