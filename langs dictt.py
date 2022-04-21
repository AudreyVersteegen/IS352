import openpyxl
wkbk = openpyxl.load_workbook("is352 language code.xlsx")
sheet = wkbk.active
langDict = {}
for line in range(2, sheet.max_row):
    code = sheet.cell(row=line, column=1).value
    lang = sheet.cell(row=line, column=2).value
    langDict[code] = lang
pprint.pprint(langDict)